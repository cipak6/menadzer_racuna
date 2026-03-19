"""
Appends invoice data to an existing Excel file.
Creates the file with headers if it doesn't exist yet.
"""
import threading
_excel_lock = threading.Lock()

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INVOICE_HEADERS = [
    'Date', 'Company', 'Company Details', 'Store Name', 'Address', 'Place',
    'Total (RSD)', 'VAT (RSD)', 'Items Count', 'Brojač računa', 'Source URL', 'Image File'
]

ITEM_HEADERS = [
    'Date', 'Company', 'Store Name', 'Item Name', 'Unit',
    'Quantity', 'Price (RSD)', 'Total (RSD)', 'VAT %', 'Source URL'
]

HEADER_FILL = PatternFill('solid', start_color='1A1A2E')
HEADER_FONT = Font(bold=True, color='E8D5B7', name='Calibri', size=11)
ALT_FILL = PatternFill('solid', start_color='F5F0E8')
BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
)


def _apply_header(ws, headers):
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _style_data_rows(ws, start_row=2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            if fill.fill_type:
                cell.fill = fill
            cell.font = Font(name='Calibri', size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center')


def _get_or_create_workbook(path: str):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    # Invoices sheet
    ws_inv = wb.active
    ws_inv.title = 'Invoices'
    _apply_header(ws_inv, INVOICE_HEADERS)
    # Items sheet
    ws_items = wb.create_sheet('Items')
    _apply_header(ws_items, ITEM_HEADERS)
    wb.save(path)
    return wb


def _is_duplicate(ws, url: str, url_col_index: int) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= url_col_index and row[url_col_index - 1] == url:
            return True
    return False


def append_invoice(excel_path: str, invoice: dict, image_filename: str = '', force: bool = False) -> dict:
    """
    Appends one invoice to the Excel file at excel_path.
    Creates the file if it doesn't exist.
    Returns {'added': bool, 'duplicate': bool, 'row': int}
    """
    with _excel_lock:
        wb = _get_or_create_workbook(excel_path)
        ws_inv = wb['Invoices']
        ws_items = wb['Items']

        # Check duplicate by URL
        url = invoice.get('url', '')
        if _is_duplicate(ws_inv, url, 11):
            if not force:
                return {'added': False, 'duplicate': True, 'row': None}
            else:
                # Delete existing row
                for row in ws_inv.iter_rows(min_row=2):
                    if row[10].value == url:
                        ws_inv.delete_rows(row[0].row)
                        break
                # Delete matching items rows
                for row in list(ws_items.iter_rows(min_row=2)):
                    if row[9].value == url:
                        ws_items.delete_rows(row[0].row)
        # --- Invoices sheet row ---
        inv_row = [
            invoice.get('date', ''),
            invoice.get('company_name', ''),
            invoice.get('company_details', ''),
            invoice.get('store_name', ''),
            f"{invoice.get('store_address', '')}, {invoice.get('store_place', '')}".strip(', '),
            invoice.get('store_municipality', ''),
            invoice.get('total_price', 0),
            invoice.get('total_vat', 0),
            len(invoice.get('items', [])),
            invoice.get('pfr_number', ''),
            url,
            image_filename,
        ]
        ws_inv.append(inv_row)

        # Format currency cells
        new_row = ws_inv.max_row
        for col in (6, 7):
            ws_inv.cell(row=new_row, column=col).number_format = '#,##0.00'

        # --- Items sheet rows ---
        for item in invoice.get('items', []):
            ws_items.append([
                invoice.get('date', ''),
                invoice.get('company_name', ''),
                invoice.get('store_name', ''),
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', 0),
                item.get('price', 0),
                item.get('total', 0),
                item.get('vat', ''),
                url,
            ])
            r = ws_items.max_row
            for col in (7, 8):
                ws_items.cell(row=r, column=col).number_format = '#,##0.00'

        _style_data_rows(ws_inv)
        _style_data_rows(ws_items)
        _auto_width(ws_inv)
        _auto_width(ws_items)

        wb.save(excel_path)
        return {'added': True, 'duplicate': False, 'row': new_row}
